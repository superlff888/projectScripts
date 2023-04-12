# !/usr/bin python3                                 
# encoding: utf-8 -*-                            
# @author: 半夏 
# @Time: 2022-11-27 13:37
# @Copyright：北京码同学
import importlib
import re
import openpyxl
from common import encry_decry
from keywords import test_keywords_run
from setting import DIR_NAME
wb = openpyxl.load_workbook(DIR_NAME + r'\data\shop_keyword.xlsx')


# 封装要执行的用例集合名称
def get_casesuitename():
    sheet_data = wb['测试用例集合']
    # 获取总行数 (行：从1开始计数)
    lines_count = sheet_data.max_row
    # 获取总列数 (列：从1开始计数)
    col_count = sheet_data.max_column
    # 用列表存储执行的测试用例集合名称
    case_suite_names = []
    for l in range(2, lines_count + 1):
        # 如果执行列是y 在添加
        if sheet_data.cell(l, 2).value == 'y':
            cell = sheet_data.cell(l, 1).value
            case_suite_names.append(cell)
    return case_suite_names


# 读取全局变量  -- 字典存储
def get_variables():
    sheet_data = wb['全局变量']
    # 获取总行数
    lines_count = sheet_data.max_row
    # 获取总列数
    cols_count = sheet_data.max_column
    # 存储格式字典
    variables = {}
    for l in range(2, lines_count + 1):
        # 第一列 -- 变量名称  -- key
        key = sheet_data.cell(l, 1).value
        # 第二列 --  变量值  -- value
        value = sheet_data.cell(l, 2).value
        variables[key] = value
    return variables


# 测试用例读取    测试用例名称相同 -- 一条测试用例(可能有n个接口)
# 外层是字典  存放测试用例  内层列表存放每条接口数据信息

def read_testcases(case_suite_name):
    """
    {'测试用例1':[
                    ['url','method','headers'],
                    ['url','method','headers']],
        '测试用例2'：[
                    ['url','method','headers'],
                    ['url','method','headers']]
    }
    """

    sheet_data = wb[case_suite_name]
    lines_count = sheet_data.max_row
    cols_count = sheet_data.max_column
    # 存放字典格式
    case_info = {}
    for l in range(2, lines_count + 1):
        # 测试用例名称
        case_name = sheet_data.cell(l, 1).value
        # 存放从url这列之后的所有数据
        line = []
        for c in range(3, cols_count + 1):
            cell = sheet_data.cell(l, c).value
            if cell is None:
                cell = ''
            line.append(cell)
        # 判断测试用例名称是否在字典中
        if case_name not in case_info:
            case_info[case_name] = [line]

        else:  # 如果存在直接追加数据
            case_info[case_name].append(line)

    return case_info


# 获取所有测试用例
def get_all_case():
    # 获取所有用例集合名称
    case_suite_names = get_casesuitename()
    test_cases = []

    for case_suite_name in case_suite_names:
        # 读取某个用例集合名称
        suite_cases = read_testcases(case_suite_name)
        # 在当前测试用例循环
        for key, value in suite_cases.items():
            cur_case = [key, value]
            test_cases.append(cur_case)
    return test_cases


# 针对excel表中动态变量 进行变量识别及替换
def regx_sub(string, var_dict):
    """

    :param string: 需要匹配的对象
    :param var_dict: 替换数据
    :return:

    动态导入
    gc = importlib.import_module("路径")  # 绝对导入（路径：工程下第一个包开始）
    反射
    getattr(gc, "demo")()

    """
    # 提取数据 ：匹配规则
    res = re.findall(r'''\$\{([A-Za-z_]+)\}''', string)  # 替换数据需要分组 ()  ，返回列表
    print(res)
    for var_name in res:
        print(var_name)
        # 公共变量字典 获取匹配到的变量值
        value = var_dict[var_name]
        # 替换value值  -- 字符串方法替换  最里面的花括号是字面量取值符号，“两个花括号折合成一个(相当于转义)， 【f'{var_name}'】等于【buyerToken】”
        string = string.replace(f'${{{var_name}}}', str(value))
        # # 使用正则替换
        # string = re.sub(r'\$\{'+var_name+r'\}', str(value), string)

    return string


# 针对excel数据存在动态函数调用，使用正则匹配并完成数据替换
def regx_exec_func(string):
    global value
    res = re.findall(r'\$\{\{(.+?)\((.+?)\)\}\}', string)
    for func in res:
        print(func)
        # 获取函数名称
        func_name = func[0]
        # 参数
        func_params = func[1]
        # python反射  判断对象是否包含对应属性 【此处非动态导入 + 反射】
        if hasattr(encry_decry, func_name):  # 参数1 加密文件 参数2 md5
            # 获取函数对象(函数内存地址)  encry_decry也可以动态导入importlib.import_module("encry_decry")
            func_method = getattr(encry_decry, func_name)
            # 执行函数传递参数
            value = func_method(func_params)
        string = string.replace(f'${{{{{func_name}({func_params})}}}}', str(value))
    return string


if __name__ == '__main__':
    # print(get_casesuitename())
    # print(get_variables())
    # pprint(read_testcases('添加购物车测试集合'))
    # print(get_all_case())
    #     string = '{"Authorization":"${buyerToken}"}'
    #     var_dict = {'buyerToken': 'aah'}
    #     print(regx_sub(string, var_dict))
        s = '''{
        "params":{
            "username":"mtx0327",
            "password":"${{md5(1234567)}}",
            "captcha":"1512",
            "uuid":"hjsdhdfhhf"
     }
    }
        '''
        print(regx_exec_func(s))

