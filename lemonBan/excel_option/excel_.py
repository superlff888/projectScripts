# @Time  : 2022/02/10 21:20
# @Author    : ╰☆H.俠ゞ
# -*-coding=utf-8-*-
# =============================================================
import openpyxl


new_wb = openpyxl.Workbook()  # 创建一个新的工作簿
"""
【新文件】
    wb = openpyxl.Workbook()  # 创建一个新的工作簿
    wb.create_sheet('sheet3')  # 出阿奴关键一个表单
    wb.save('test.xlsx')  # 将工作簿保存为文件。如果保存的文件存在，将会覆盖原来的文件
【已存在的文件】
    openpyxl.load_workbook(path)  # 打开工作簿
"""

# _wb = openpyxl.load_workbook('../data/data.xlsx')  # 打开一个已存在的工作簿
_wb = openpyxl.load_workbook(r'D:\Program Files\pythonProject_CTTQ\lemonBan\data\data.xlsx')  # 打开一个已存在的工作簿
# 获取所有的sheet表单
print(_wb.sheetnames)
# 选中表单，获取操作的sheet表单
sheet = _wb["Sheet1"]

"""
【读取数据】
"""
'''
指定单元格读取
'''
# 根据行、列获取到表格
cell_ = sheet.cell(row=1, column=2)  # row:行  column:列
# 获取表格中的数据
print(f"第一行，第二列的数据为：{cell_.value}")


'''
按行、列读取数据
'''
rows_data = sheet.rows  # 获取到行
print(f"行数据: {list(rows_data)}")  # 以列表形式打印行数据

column_data = sheet.columns  # 获取到列
print(list(column_data))  # 以列表形式打印列数据

'''
获取最大行、获取最大列(以有数据数据的那一行和列为准，)
'''
max_row = sheet.max_row
max_column = sheet.max_column
print(f'最大行数：{max_row}')
print(f'最大列数：{max_column}')

'''
获取第二列、第三列的所有数据
'''
list_data = []
for i in range(2, max_row+1):
    list_row = []
    for j in range(2, 4):  # 需要取3列、4列数据
        value = sheet.cell(i, j).value
        list_row.append(value)  # 用多个列表分别存放每行的数据
    list_data.append(list_row)  # 行列表数据作为一个元素保存在另外一个大的列表中
print(list_data)

"""
【写入数据】

"""
sheet.cell(row=1, column=1, value='Num')
# 保存数据到文件(相当于点击保存)；注意：修改文件数据时，文件一定要关闭状态
_wb.save('test_.xlsx')  # 单元格有数据，修改原来的数据为新数据

"""
创建sheet表单
"""
# _wb.create_sheet('test_case1')
# _wb.save(r'../data/data.xlsx')

'''
删除表单(要先获取操作的sheet表单，即先选中该表单)
'''
# sheet2 = _wb["Sheet2"]
# _wb.remove(sheet2)  # 参数应该是一个对象
# _wb.save(r'../data/data.xlsx')


"""
关闭文件
"""
_wb.close()

"""
del 删除目标工作簿或sheet表单
"""
# del _wb['Sheet3']  # 不需要先获取相应表单或工作簿
# del _wb  # 不需要先获取相应表单或工作簿