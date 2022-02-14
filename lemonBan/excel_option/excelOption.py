# @Time  : 2021/06/18 13:09
# @Author    : ╰☆H.俠ゞ
# -*-coding=utf-8-*-
# =============================================================

import openpyxl

"""
封装一个读取excel表格的类
    1、实现读取用例数据
    2、实现写入数据
"""


class ExcelOption:
    def __init__(self, filename, sheetname):
        """
        初始化读取对象
        :param filename: 文件路径 -> str
        :param sheetname: sheet表单名 -> str
        """
        self.wb_ = openpyxl.load_workbook(filename)  # 打开工作簿
        self.sheet_ = self.wb_[sheetname]  # 选中指定sheet表单

    def read_excel(self, flag_: str = None, num=None,
                   column=None):  # 可以将此方法改成静态方法@staticmethod，此时方法中就没有用到self(故可不写self)
        """
        可以将列做成参数
        :param num: 起始列
        :param column:
        :param flag_:
        :return:
        """
        if flag_ == 'max':
            max_row = self.sheet_.max_row
            max_column = self.sheet_.max_column
            list_data = []
            for i in range(2, max_row + 1):  # 忽略表头
                list_row = []
                for j in range(1, max_column + 1):  # 取所有列的话，用 max_column+1
                    value = self.sheet_.cell(i, j).value
                    list_row.append(value)  # 用多个列表分别存放每行的数据
                list_data.append(list_row)  # 行列表数据作为一个元素保存在另外一个大的列表中
            return list_data

        if flag_ is None:
            max_row = self.sheet_.max_row
            list_data = []
            for i in range(2, max_row + 1):
                list_row = []
                for j in range(num, column + 1):  # 需要取2列、3列数据 ; 取所有列的话，用 max_column+1
                    value = self.sheet_.cell(i, j).value
                    list_row.append(value)  # 用多个列表分别存放每行的数据
                list_data.append(list_row)  # 行列表数据作为一个元素保存在另外一个大的列表中
            return list_data

    # 按行读取
    def readlines_excel(self):
        pass

    def write_excel(self):
        pass

# excel_data = ExcelOption.read_excel()
