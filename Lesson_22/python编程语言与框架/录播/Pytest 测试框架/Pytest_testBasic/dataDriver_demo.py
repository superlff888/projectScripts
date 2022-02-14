# @Time  : 2022/01/23 21:31
# @Author    : ╰☆H.俠ゞ
# -*-coding=utf-8-*-
# =============================================================

import openpyxl

# 获取工作簿
book = openpyxl.load_workbook('openpyxl.xlsx')

# 读取工作表
sheets = book.sheetnames  # 获取全部sheet页
sheet = book.active
print(sheets[0])
print(sheet)

# 读取单个单元格
cell_a1 = sheet['A1']
print(cell_a1.value)
# cell_a3 = sheet.cell(column=1, row=3)  # A3

# 读取多个连续单元格
cells = sheet["A1":"C3"]
print(cells)
# # 获取单元格的值
# var_a = cell_a1.value
